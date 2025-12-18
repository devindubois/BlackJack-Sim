from bj import Game, PlayerHand, DealerHand, Deck, Card
import pandas as pd
from openpyxl import load_workbook




class AutoGame:
    def count_cards(hands):
        """Return Hi-Lo count for a collection of hands.

        Rules: 2-6 => +1, 10/J/Q/K/A => -1, all others => 0.
        `hands` may be an iterable of PlayerHand objects or iterables of card-like objects.
        """
        total = 0
        for hand in hands:
            # support both PlayerHand objects (with .cards) and plain iterables
            cards = getattr(hand, 'cards', hand)
            for c in cards:
                rank = getattr(c, 'rank', c)
                r = str(rank)
                if r in ('2', '3', '4', '5', '6'):
                    total += 1
                elif r in ('10', 'J', 'Q', 'K', 'A'):
                    total -= 1
        return total

    def determine_strategy(card_count: int) -> str:
        # For now return a static path
        if card_count >= 5:
            return "base_strategy.xlsx"
        elif card_count >= 5:
            return "base_strategy.xlsx"
        elif card_count <= -5:
            return "base_strategy.xlsx"
        else:
            return "base_strategy.xlsx"
        
    def determine_bet_multiple(true_card_count: int) -> str:
        # For now return a static path
        if true_card_count >= 10:
            return 10
        if true_card_count >= 5:
            return 4
        elif true_card_count >= 2:
            return 2
        elif true_card_count >= 0:
            return 1
        elif true_card_count >= -5:
            return 1
        else:
            return 0

    class Strategy:
        def __init__(self, path: str):
            self.hard_df, self.soft_df, self.split_df = self.read_exact_ranges(path)

        def read_exact_ranges(self, path: str):
            wb = load_workbook(path, data_only=True)
            ws = wb.active

            def get_range(ws, ref):
                data = ws[ref]
                rows = [[cell.value for cell in row] for row in data]
                return pd.DataFrame(rows)

            hard = get_range(ws, "B3:K20")
            soft = get_range(ws, "B23:K32")
            split = get_range(ws, "B35:K44")
            if(False):
                print("Loaded Strategy:")
                print("Hard Hands:")
                print(hard)
                print("Soft Hands:")
                print(soft)
                print("Split Hands:")
                print(split)
            return hard, soft, split

        def get_action(self, player_hand: PlayerHand, dealer_value, splitallowed=True) -> str:

            if player_hand.can_split() and splitallowed:
                row = player_hand.cards[0].get_value() - 2
                col = dealer_value - 2
                if self.split_df.iat[row, col] == 'Y':
                    return 'v'
            hand_value = player_hand.get_value()
            if player_hand.has_soft_ace():
                row = hand_value - 12  # Soft hands start from 12
                col = dealer_value - 2
                action = self.soft_df.iat[row, col]
                if action == "DS":
                    if player_hand.num_cards() == 2:
                        return 'd' 
                    else:
                        return 's'
                else:
                    if action == "D":
                        if player_hand.num_cards() == 2:
                            return 'd' 
                        else:
                            return 'h'
                    return action.lower()
            else:
                row = hand_value - 4  # Hard hands start from 4
                col = dealer_value - 2
                action = self.hard_df.iat[row, col]
                if action == "R":
                    action = "H"  
                if action == "D" and player_hand.num_cards() != 2:
                    action = "H"
                return action.lower()

    def auto_play_loop(bet_amount=1, num_games=100, balance=1000, num_decks=8):
        deck = Deck(num_decks)
        deck.shuffle()
        total_profit = 0
        shoe_profit = 0
        card_count = 0
        open('results.txt', 'w').close()  # Clear results file at start
        
        for _ in range(num_games):
            modified_bet_amount = bet_amount * AutoGame.determine_bet_multiple(card_count/num_decks)
            if balance < modified_bet_amount:
                print("Out of money! Time to go home.")
                break
            if modified_bet_amount == 0:
                print("Time to leave the table, count is too low. Shoe Profit: ", shoe_profit)
                deck = Deck(num_decks=8)
                deck.shuffle()
                card_count = 0
                shoe_profit = 0
            true_card_count = card_count/num_decks
            strategy_path = AutoGame.determine_strategy(true_card_count)
            strategy = AutoGame.Strategy(strategy_path)
            game = deck.new_game()
            round_result = AutoGame.played_hand(game, modified_bet_amount, balance, strategy)
            profit = Game.interpret_result(round_result)
            card_count += Game.interpret_card_count(round_result)
            balance += profit
            total_profit += profit
            shoe_profit += profit
            with open('results.txt', 'a') as f:
                f.write(f"hand{_}: balance: {balance} card count: {card_count} \"True\" card count: {card_count/num_decks}\n")
            game.end_game()
            #print("hand", _, "result:", round_result, "balance:", balance)
            if len(deck.cards) < (52 * 8 * 0.25):  # Less than 25% of cards remain
                    print("Reshuffling deck. Shoe Profit: ", shoe_profit)
                    deck = Deck(num_decks=8)
                    deck.shuffle()
                    card_count = 0
                    shoe_profit = 0

    


    def played_hand_split(game, bet_amount, balance, strategy, split_hand, handnum):
        """
        Play ONLY the player's actions for a single split hand and return (final_hand, bet_amount).
        No dealer logic here. Robust to different deal_split implementations.
        """
        # Determine target index robustly
        target_idx = handnum
        if split_hand is not None:
            pre_len = len(game.player_hands)
            game.deal_split(split_hand)
            post_len = len(game.player_hands)

            if post_len > pre_len:
                # A new hand was appended; play that new hand
                target_idx = post_len - 1
            else:
                # No new hand created. Constrain to valid index.
                target_idx = min(handnum, post_len - 1)

        # Constrain index again for safety
        if not game.player_hands:
            raise RuntimeError("No player hands available after split.")
        target_idx = max(0, min(target_idx, len(game.player_hands) - 1))

        hand = game.player_hands[target_idx]

        while not hand.is_busted():
            # Auto-stop at 21
            if hand.get_value() == 21:
                break

            action = strategy.get_action(hand, game.dealer_hand.get_card_shown_value(), splitallowed=False)

            if action == 'h':
                game.player_hit(handnum=target_idx)
                hand = game.player_hands[target_idx]
                if hand.is_busted():
                    break
                continue

            elif action == 's':
                break

            elif action == 'd' and hand.num_cards() == 2 and balance >= bet_amount:
                game.player_hit(handnum=target_idx)
                bet_amount *= 2
                hand = game.player_hands[target_idx]
                # After a double, stand automatically
                break

            #else:
                #print("Invalid action.")

        return (game.player_hands[target_idx], bet_amount)




    def played_hand(game, bet_amount, balance, strategy):
        """
        Main hand flow. Uses game.player_hands as a list where the first hand is index 0.
        """
        # Initialize a fresh hand slot before dealing
        game.new_hand()

        dealer_hand = game.dealer_hand
        game.deal_initial()
        player_hand = game.player_hands[0]
        if player_hand.blackjack():
            #dealer_hand.show_cards()
            if dealer_hand.blackjack():
                #print("Push!")
                return ("P", bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))
            #print("Blackjack! You win!")
            return ("W!", bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))

        if dealer_hand.blackjack():
            return ("L", bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))
        
        while not player_hand.is_busted():
            
            #determine action
            if player_hand.get_value() == 21:
                #print("21: no more action")
                action = 's'
            else:
                action = strategy.get_action(player_hand, dealer_hand.get_card_shown_value(), splitallowed=True)

            if action == 'h':
                game.player_hit(handnum=0)
                player_hand = game.player_hands[0]
                if player_hand.is_busted():
                    #print("Player busted!")
                    #dealer_hand.show_cards()
                    return ("L", bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))
                continue

            elif action == 's':
                game.dealer_play()
                #dealer_hand.show_cards()
                result = game.get_winner(0)
                #print(result)
                return (result, bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))

            elif action == 'd' and player_hand.num_cards() == 2 and balance >= bet_amount:
                bet_amount *= 2
                game.player_hit(handnum=0)
                player_hand = game.player_hands[0]
                if player_hand.is_busted():
                    #print("Player busted after doubling down!")
                    #dealer_hand.show_cards()
                    return ("L", bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))
                game.dealer_play()
                #dealer_hand.show_cards()
                result = game.get_winner(0)
                #print(result)
                return (result, bet_amount, AutoGame.count_cards([player_hand, dealer_hand]))

            elif action == 'v' and player_hand.can_split() and balance >= bet_amount:
                #print("splitting")

                # Capture the two seed cards before split mutates the hand
                seed_left = player_hand.cards[0]
                seed_right = player_hand.cards[1]
                game.new_hand()
                # Play both split hands
                #print("1st split hand:")
                h1, b1 = AutoGame.played_hand_split(game, bet_amount, balance=balance-bet_amount, strategy=strategy, split_hand=seed_left, handnum=0)

                #print("2nd split hand:")
                h2, b2 = AutoGame.played_hand_split(game, bet_amount, balance=balance-bet_amount, strategy=strategy, split_hand=seed_right, handnum=1)

                # Dealer plays once for both hands
                game.dealer_play()
                #dealer_hand.show_cards()

                # Score each hand independently
                game.player_hands[0] = h1
                game.player_hands[1] = h2
                r1 = game.get_winner(0)
                r2 = game.get_winner(1)
                #print(r1)
                #print(r2)
                return [(r1, b1, AutoGame.count_cards([h1, h2, dealer_hand])), (r2, b2, 0)]
            #else:

                #print("Invalid action: " + action)
                #print("Hand value:", player_hand.get_value())``
                #print("Dealer showing:", dealer_hand.get_card_shown_value())
                

        return ("E", bet_amount, 0)
    

    
